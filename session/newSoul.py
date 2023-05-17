""" ================================================== 
 Description     : take 3
 Author          : Suyog Chavan
 Reach me here   : https://suyogschavan.github.io 
 Date of creation: Friday, 14 Apr 2023 
 Copyright       : ©suyogschavan03@gmail.com
================================================== """


import random
from datetime import datetime, timedelta
import openpyxl

# List of Indian company names
company_names = ['Tata Consultancy Services', 'Reliance Industries', 'HDFC Bank', 'Infosys', 'Wipro', 'Bharti Airtel', 'Coal India', 'Indian Oil', 'Larsen & Toubro', 'Hindustan Unilever']

# List of Indian names
indian_names = ['Aarav', 'Aditi', 'Advait', 'Amitabh', 'Arun', 'Deepika', 'Gaurav', 'Isha', 'Kavya', 'Mihir', 'Neha', 'Pranav', 'Raj', 'Rani', 'Sanjay', 'Shalini']

# Possible transaction types
transaction_types = ['Bank Transfer', 'Money Transfer', 'College Fees', 'Hospital Fees']

# Start and end dates for transactions
start_date = datetime(2022, 1, 1)
end_date = datetime(2022, 12, 31)

# Create a workbook and sheet
workbook = openpyxl.Workbook()
sheet = workbook.active

# Write headers
sheet['A1'] = 'Time'
sheet['B1'] = 'Amount'
sheet['C1'] = 'Transaction Type'
sheet['D1'] = 'Payee Info'
sheet['E1'] = 'Receiver Info'

# Set starting balance
balance = 50000

# Generate transactions
for i in range(2, 1002):
    # Random amount
    if i % 8 == 0:
        amount = random.randint(10000, 20000)
    else:
        amount = random.randint(100, 2000)

    # Random transaction type
    if i % 6 == 0:
        transaction_type = 'College Fees'
    elif i % 8 == 0:
        transaction_type = 'Hospital Fees'
    else:
        transaction_type = 'Bank Transfer'

    # Random payee
    if i % 5 == 0:
        payee = random.choice(company_names)
    else:
        payee = random.choice(indian_names)

    # Random receiver
    if i % 10 == 0:
        receiver = random.choice(company_names)
    else:
        receiver = random.choice(indian_names)

    # Calculate remaining balance
    if transaction_type != 'Hospital Fees':
        balance -= amount

    # Write transaction to sheet
    sheet.cell(row=i, column=1).value = start_date + timedelta(minutes=i*10)
    sheet.cell(row=i, column=2).value = amount
    sheet.cell(row=i, column=3).value = transaction_type
    sheet.cell(row=i, column=4).value = payee
    sheet.cell(row=i, column=5).value = receiver
    sheet.cell(row=i, column=6).value = balance

# Save workbook
workbook.save('transactions.xlsx')
